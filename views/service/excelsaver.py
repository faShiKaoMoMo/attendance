import io
import smtplib
from datetime import datetime, timedelta
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import openpyxl
from openpyxl.styles import Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def create_combined_workbook(data, start_date, end_date):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "考勤统计报表"

    # --- 样式定义 ---
    center_align = Alignment(horizontal='center', vertical='center')

    # 隔行变色：淡黄色
    light_yellow_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")

    # 边框样式
    thin_grey_side = Side(style='thin', color='D3D3D3')
    thin_border = Border(left=thin_grey_side, right=thin_grey_side,
                         top=thin_grey_side, bottom=thin_grey_side)

    # --- 准备表头 ---
    dates = []
    current_date = datetime.strptime(start_date, "%Y%m%d")
    end_dt = datetime.strptime(end_date, "%Y%m%d")
    while current_date <= end_dt:
        dates.append(current_date.strftime("%m-%d"))
        current_date += timedelta(days=1)

    # 左侧固定列 (包含新增的“出差补录时长”和“最终出勤时长”)
    summary_headers = [
        "排名",
        "姓名",
        "周期出勤时长",
        "必在时段不达标次数",
        "日均出勤时长",
        "是否达标"
    ]
    headers = summary_headers + ["项目"] + dates

    ws.append(headers)

    # --- 设置列宽 ---
    # 根据新增列调整宽度映射
    ws.column_dimensions['A'].width = 6  # 排名
    ws.column_dimensions['B'].width = 12  # 姓名
    ws.column_dimensions['C'].width = 18  # 周期出勤时长（新列）
    ws.column_dimensions['D'].width = 20  # 必在时段不达标次数
    ws.column_dimensions['E'].width = 15  # 日均出勤时长
    ws.column_dimensions['F'].width = 10  # 是否达标
    ws.column_dimensions['G'].width = 12  # 项目

    # 动态日期列的宽度设置
    for i in range(len(dates)):
        # 注意：这里基于 headers 的固定列长度自动计算起始列，无需手动修改
        col_letter = get_column_letter(len(summary_headers) + 2 + i)
        ws.column_dimensions[col_letter].width = 20

    # 设置表头样式
    ws.row_dimensions[1].height = 25
    for cell in ws[1]:
        cell.alignment = center_align
        cell.border = thin_border
        cell.font = openpyxl.styles.Font(bold=True)

    # --- 填充数据 ---
    row_items = [
        ('checkin_times', '打卡时间'),
        ('class_times', '上课时间'),
        ('leave_times', '请假时间'),
        ('duration', '统计时长')
    ]

    current_row = 2
    for person_index, person in enumerate(data):
        start_row_for_person = current_row

        # 1. 获取数据
        # 注意：data 中的 '日均考勤时长' 已经在上一步计算逻辑中更新为基于最终时长的均值
        avg_hours = person.get("日均考勤时长", 0)
        is_target_met = "是" if avg_hours >= 8 and person.get("必在时段不达标次数", 0) <= 4 else "否"

        # 构造左侧固定列数据，加入新增字段
        actual_hours = person.get("实际出勤时长", 0)
        travel_hours = person.get("出差补录时长", 0)
        final_hours = person.get("最终出勤时长", 0)
        # 构造周期出勤时长的文本内容
        total_hours_text = f"{final_hours}\n（{actual_hours} + {travel_hours}）"
        static_data = [
            person.get("排名", ""),
            person.get("姓名", ""),
            total_hours_text,  # ⬅ 新的合并列内容
            person.get("必在时段不达标次数", 0),
            avg_hours,
            is_target_met
        ]

        # 2. 写入左侧固定列并合并单元格
        for i, value in enumerate(static_data, start=1):
            ws.cell(row=current_row, column=i, value=value)
            # 合并4行
            ws.merge_cells(start_row=current_row, start_column=i,
                           end_row=current_row + 3, end_column=i)

        item_col = len(summary_headers) + 1
        for row_offset, (_, label) in enumerate(row_items):
            ws.cell(row=current_row + row_offset, column=item_col, value=label)

        # 3. 写入右侧日期详情
        for col_offset, day_key in enumerate(dates):
            current_col = len(summary_headers) + 2 + col_offset
            day_times = person.get(day_key, {})

            for row_offset, (key, _) in enumerate(row_items):
                time_value = day_times.get(key, '-')
                ws.cell(row=current_row + row_offset, column=current_col, value=time_value)

        # 4. 统一应用样式 (边框 + 对齐 + 隔人换色)
        for row_index in range(start_row_for_person, start_row_for_person + 4):
            max_line_count = 1
            for col_index in range(1, len(headers) + 1):
                value = ws.cell(row=row_index, column=col_index).value
                if isinstance(value, str):
                    max_line_count = max(max_line_count, value.count("\n") + 1)
            ws.row_dimensions[row_index].height = max(24, min(120, 18 * max_line_count))

            for col_index in range(1, len(headers) + 1):
                cell = ws.cell(row=row_index, column=col_index)

                # 这里要加 wrap_text=True 才能让 \n 生效
                cell.alignment = Alignment(
                    horizontal='center',
                    vertical='center',
                    wrap_text=True
                )

                cell.border = thin_border

                if person_index % 2 == 0:
                    cell.fill = light_yellow_fill

        current_row += 4

    return wb


def send_email_with_attachments(recipient_list, subject, body, attachments):
    sender_email = 'xmulab309@163.com'
    sender_password = 'YSRQQMBRTDNEZNZI'
    smtp_server = 'smtp.163.com'
    smtp_port = 465

    msg = MIMEMultipart()
    msg['Subject'] = subject
    msg['From'] = sender_email
    msg['To'] = ', '.join(recipient_list)
    msg.attach(MIMEText(body, 'plain', 'utf-8'))

    for attachment in attachments:
        filename = attachment['filename']
        content = attachment['content']

        part = MIMEApplication(content)
        part.add_header('Content-Disposition', 'attachment', filename=filename)
        msg.attach(part)

    try:
        client = smtplib.SMTP_SSL(smtp_server, smtp_port)
        client.login(sender_email, sender_password)
        client.sendmail(sender_email, recipient_list, msg.as_string())
        client.quit()
        return True
    except smtplib.SMTPException as e:
        print(f"邮件发送失败: {e}")
        raise e


def excel(conn, cursor, data, req_data):
    start_date = req_data['start_date']
    end_date = req_data['end_date']

    # 生成唯一的 Excel 文件
    # data 中已经包含了 "出差补录时长" 和 "最终出勤时长" 字段
    combined_wb = create_combined_workbook(data, start_date, end_date)

    # 将 Excel 保存到内存
    virtual_workbook = io.BytesIO()
    combined_wb.save(virtual_workbook)
    virtual_workbook.seek(0)

    return virtual_workbook.getvalue()
